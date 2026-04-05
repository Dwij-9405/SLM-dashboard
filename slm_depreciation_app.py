import streamlit as st
import pandas as pd
import numpy as np
import openpyxl as px 
from io import BytesIO
from datetime import date, datetime
import json
import hashlib
import os

st.set_page_config(page_title="SLM Depreciation Calculator", layout="wide")

# ── File-based user registry ──────────────────────────────────────────────────
USERS_FILE   = "slm_users.json"
HISTORY_FILE = "slm_history.json"

def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def load_users() -> dict:
    """Load {username: hashed_password} from file. Seed admin if file missing."""
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    # First-run seed: one admin account so the app is never completely empty
    seed = {"admin": hash_password("admin123")}
    with open(USERS_FILE, "w") as f:
        json.dump(seed, f, indent=2)
    return seed

def save_users(users: dict):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)

def user_exists(username: str) -> bool:
    return username.strip().lower() in {k.lower() for k in load_users()}

def register_user(username: str, password: str) -> tuple[bool, str]:
    """Returns (success, message)."""
    username = username.strip()
    if not username:
        return False, "Username cannot be empty."
    if len(username) < 3:
        return False, "Username must be at least 3 characters."
    if len(password) < 6:
        return False, "Password must be at least 6 characters."
    users = load_users()
    if username.lower() in {k.lower() for k in users}:
        return False, f"Username **{username}** is already taken. Please choose another."
    users[username] = hash_password(password)
    save_users(users)
    return True, f"Account created for **{username}**. You can now sign in."

def check_login(username: str, password: str) -> bool:
    users = load_users()
    stored = users.get(username) or users.get(username.strip())
    return stored == hash_password(password)

# ── History helpers ───────────────────────────────────────────────────────────
def load_history() -> list:
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_history(history: list):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2, default=str)

def add_history_entry(client, start_dt, end_dt, num_rows, totals: dict):
    history = load_history()
    entry = {
        "timestamp":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "client":         client or "—",
        "period_start":   start_dt.strftime("%d/%m/%Y"),
        "period_end":     end_dt.strftime("%d/%m/%Y"),
        "num_assets":     num_rows,
        "total_cost":     totals.get("cost", 0),
        "total_depr":     totals.get("depr", 0),
        "total_closing":  totals.get("closing", 0),
        "total_accum":    totals.get("accum", 0),
    }
    history.insert(0, entry)
    history = history[:50]
    save_history(history)

# ── Auth screen (Sign In + Create Account) ────────────────────────────────────
def auth_screen():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;700&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
        html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    col_l, col_m, col_r = st.columns([1, 2, 1])
    with col_m:
        st.markdown("## 📊 SLM Depreciation Calculator")
        st.markdown("---")

        # Tab toggle stored in session state
        if "auth_tab" not in st.session_state:
            st.session_state["auth_tab"] = "signin"

        tab_signin, tab_signup = st.tabs(["🔑 Sign In", "🆕 Create Account"])

        # ── Sign In ──────────────────────────────────────────────────────────
        with tab_signin:
            st.markdown("#### Welcome back")
            with st.form("login_form"):
                username  = st.text_input("Username", placeholder="Enter your username")
                password  = st.text_input("Password", type="password", placeholder="Enter your password")
                submitted = st.form_submit_button("Sign In", use_container_width=True, type="primary")
                if submitted:
                    if not username or not password:
                        st.error("Please enter both username and password.")
                    elif check_login(username, password):
                        st.session_state["logged_in"] = True
                        st.session_state["username"]  = username.strip()
                        st.rerun()
                    else:
                        st.error("❌ Invalid username or password.")
            st.caption("New user? Switch to the **Create Account** tab →")

        # ── Create Account ───────────────────────────────────────────────────
        with tab_signup:
            st.markdown("#### Create your account")
            with st.form("signup_form"):
                new_username  = st.text_input("Choose a Username", placeholder="Min. 3 characters")
                new_password  = st.text_input("Choose a Password", type="password", placeholder="Min. 6 characters")
                confirm_pass  = st.text_input("Confirm Password",  type="password", placeholder="Re-enter password")
                reg_submitted = st.form_submit_button("Create Account", use_container_width=True, type="primary")
                if reg_submitted:
                    if not new_username or not new_password or not confirm_pass:
                        st.error("All fields are required.")
                    elif new_password != confirm_pass:
                        st.error("❌ Passwords do not match.")
                    else:
                        ok, msg = register_user(new_username, new_password)
                        if ok:
                            st.success(f"✅ {msg}")
                            st.info("Switch to the **Sign In** tab to log in.")
                        else:
                            st.error(f"❌ {msg}")
            st.caption("Already have an account? Switch to the **Sign In** tab ←")

# ── Session state init ────────────────────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    auth_screen()
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APP  (only reached when logged in)
# ══════════════════════════════════════════════════════════════════════════════

# ── Column mapping (A-O) ─────────────────────────────────────────────────────
COLS = {
    "A": "Asset class",
    "B": "Asset",
    "C": "Sub No.",
    "D": "Capitalized date",
    "E": "Cost",
    "F": "Salvage Value",
    "G": "Useful life",
    "H": "Book Value as on Start Date",
    "I": "sale date",
    "J": "Depre as per client",
    "K": "Depreciation",
    "L": "Closing Value",
    "M": "Profit / Loss",
    "N": "Sale Value",
    "O": "Accumulated Depreciation",
}

REQUIRED_INPUT_COLS = ["A","B","C","D","E","G","I","J","N"]
COMPUTED_COLS       = ["F","H","K","L","M","O"]
ALL_COLS_ORDERED    = list(COLS.keys())

# Asset classes that carry ZERO depreciation
ZERO_DEPR_ASSET_CLASSES = {"freehold land", "lease hold land", "leasehold land"}

# ── Header / nav ──────────────────────────────────────────────────────────────
hcol1, hcol2 = st.columns([6, 1])
with hcol1:
    st.title("📊 SLM Depreciation Calculator")
    st.markdown("Straight Line Method | Pro-rata Depreciation")
with hcol2:
    st.markdown(f"<br>👤 **{st.session_state['username']}**", unsafe_allow_html=True)
    if st.button("🚪 Logout"):
        st.session_state["logged_in"] = False
        st.session_state["username"]  = ""
        st.rerun()

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_calc, tab_hist = st.tabs(["🧮 Calculator", "🕓 History"])

# ── VBA-matching month calculation ───────────────────────────────────────────
def date_diff_months_vba(d1: date, d2: date) -> float:
    whole_months = (d2.year - d1.year) * 12 + (d2.month - d1.month)
    fractional   = (d2.day - d1.day) / 30.0
    return whole_months + fractional

# ── Core calculation (VBA-aligned) ───────────────────────────────────────────
def calculate_slm(df_in, start_dt: date, end_dt: date):
    df = df_in.copy()

    for col in ALL_COLS_ORDERED:
        if col not in df.columns:
            df[col] = np.nan

    for col in ["D", "I"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    for idx, row in df.iterrows():
        asset_class  = str(row.get("A", "")).strip()
        asset_class_lower = asset_class.lower()
        cost         = float(row.get("E", 0) or 0)
        useful_life  = float(row.get("G", 0) or 0)
        cap_date_raw = row.get("D")
        sale_val     = float(row.get("N", 0) or 0)

        cap_date  = cap_date_raw.date() if pd.notna(cap_date_raw) else start_dt

        sale_date_raw = row.get("I")
        sale_date = sale_date_raw.date() if pd.notna(sale_date_raw) else None

        # ── Zero depreciation for land asset classes ─────────────────────────
        if asset_class_lower in ZERO_DEPR_ASSET_CLASSES:
            book_value    = cost
            salvage       = 0.0
            period_depr   = 0.0
            accum_depr    = 0.0
            closing_value = cost
            profit_loss   = 0.0
            if sale_date and start_dt <= sale_date <= end_dt:
                profit_loss = sale_val - closing_value
            df.at[idx, "F"] = round(salvage, 2)
            df.at[idx, "H"] = round(book_value, 2)
            df.at[idx, "K"] = round(period_depr, 2)
            df.at[idx, "L"] = round(closing_value, 2)
            df.at[idx, "M"] = round(profit_loss, 2)
            df.at[idx, "O"] = round(accum_depr, 2)
            continue

        # ── F: Salvage Value ─────────────────────────────────────────────────
        if asset_class_lower == "software":
            salvage = 0.0
        else:
            salvage = cost * 0.05

        # ── Annual / Monthly depreciation ────────────────────────────────────
        if useful_life > 0:
            annual_depr  = (cost - salvage) / useful_life
            monthly_depr = annual_depr / 12.0
        else:
            annual_depr  = 0.0
            monthly_depr = 0.0

        # ── calcStart / calcEnd ──────────────────────────────────────────────
        calc_start = max(cap_date, start_dt)
        calc_end   = min(sale_date, end_dt) if sale_date else end_dt

        # ── K: Period Depreciation ───────────────────────────────────────────
        if calc_end >= calc_start and cap_date <= end_dt:
            total_months = date_diff_months_vba(calc_start, calc_end)
            total_months = max(total_months, 0.0)
            period_depr  = monthly_depr * total_months
        else:
            period_depr  = 0.0

        # ── O: Accumulated Depreciation ──────────────────────────────────────
        if calc_end >= cap_date:
            total_months_acc = date_diff_months_vba(cap_date, calc_end)
            total_months_acc = max(total_months_acc, 0.0)
            accum_depr       = monthly_depr * total_months_acc
        else:
            accum_depr = 0.0

        # ── Salvage control ──────────────────────────────────────────────────
        max_possible_depr = cost - salvage
        if accum_depr > max_possible_depr:
            prior_depr = accum_depr - period_depr
            if prior_depr >= max_possible_depr:
                period_depr = 0.0
            else:
                period_depr = max_possible_depr - prior_depr
            accum_depr = max_possible_depr

        # ── L: Closing Value ─────────────────────────────────────────────────
        closing_value = cost - accum_depr

        # ── H: Book Value as on Start Date ───────────────────────────────────
        if cap_date < start_dt and monthly_depr > 0:
            months_to_start   = date_diff_months_vba(cap_date, start_dt)
            months_to_start   = max(months_to_start, 0.0)
            acc_before_period = monthly_depr * months_to_start
            acc_before_period = min(acc_before_period, max_possible_depr)
            book_value        = cost - acc_before_period
        else:
            book_value = cost

        # ── M: Profit / Loss ─────────────────────────────────────────────────
        if sale_date and start_dt <= sale_date <= end_dt:
            profit_loss = sale_val - closing_value
        else:
            profit_loss = 0.0

        df.at[idx, "F"] = round(salvage, 2)
        df.at[idx, "H"] = round(book_value, 2)
        df.at[idx, "K"] = round(period_depr, 2)
        df.at[idx, "L"] = round(closing_value, 2)
        df.at[idx, "M"] = round(profit_loss, 2)
        df.at[idx, "O"] = round(accum_depr, 2)

    existing_ordered = [c for c in ALL_COLS_ORDERED if c in df.columns]
    return df[existing_ordered]

# ── Map Excel columns to internal keys ───────────────────────────────────────
def normalise_columns(df):
    letter_map = {v: k for k, v in COLS.items()}
    rename = {}
    for col in df.columns:
        col_s = str(col).strip()
        if col_s in COLS:
            rename[col] = col_s
        elif col_s in letter_map:
            rename[col] = letter_map[col_s]
    df = df.rename(columns=rename)
    return df

# ── Build display dataframe ───────────────────────────────────────────────────
def display_df(df):
    ordered = [c for c in ALL_COLS_ORDERED if c in df.columns]
    show    = df[ordered].copy()
    show    = show.rename(columns=COLS)
    for c in ["Capitalized date", "sale date"]:
        if c in show.columns:
            show[c] = pd.to_datetime(show[c], errors="coerce").dt.strftime("%d/%m/%Y")
    return show

# ── Export to Excel ───────────────────────────────────────────────────────────
def to_excel(df, client, s_date, e_date):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Shared style helpers ──────────────────────────────────────────────────
    thin        = Side(style="thin",   color="AAAAAA")
    thick       = Side(style="medium", color="1F4E79")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    outer_border= Border(left=thick, right=thick, top=thick, bottom=thick)
    header_fill = PatternFill("solid", start_color="1F4E79")
    calc_fill   = PatternFill("solid", start_color="2E75B6")
    data_calc_fill = PatternFill("solid", start_color="D6E4F0")
    total_fill  = PatternFill("solid", start_color="BDD7EE")
    title_fill  = PatternFill("solid", start_color="0A2F5C")   # deep navy for title band
    meta_fill   = PatternFill("solid", start_color="EBF3FB")   # light blue for meta area
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_font  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    meta_label_font = Font(bold=True, color="1F4E79", size=10)
    meta_value_font = Font(color="1F3050", size=10)
    NUM_FMT  = '#,##0.00'
    DATE_FMT = "DD/MM/YYYY"

    def apply_outer_border(ws, min_row, max_row, min_col, max_col):
        """Draw a thick outer border around a rectangular range."""
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for cell in row:
                l = thick if cell.column == min_col else thin
                r = thick if cell.column == max_col else thin
                t = thick if cell.row    == min_row else thin
                b = thick if cell.row    == max_row else thin
                cell.border = Border(left=l, right=r, top=t, bottom=b)

    # ════════════════════════════════════════════════════════════════════════
    #  SHEET 1 – SLM Depreciation (detail)
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "SLM Depreciation"

    ordered_keys = [c for c in ALL_COLS_ORDERED if c in df.columns]
    n_cols       = len(ordered_keys)

    # ── Row 1: Full-width title banner ───────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="STRAIGHT LINE METHOD (SLM) DEPRECIATION SCHEDULE")
    title_cell.font      = title_font
    title_cell.fill      = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Client Name (label in col1, value spanning rest) ─────────────
    ws.cell(row=2, column=1, value="Client Name").font      = meta_label_font
    ws.cell(row=2, column=1).fill                           = meta_fill
    ws.cell(row=2, column=1).alignment                      = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_cols)
    cn_cell = ws.cell(row=2, column=2, value=client)
    cn_cell.font      = Font(bold=True, color="1F3050", size=11)
    cn_cell.fill      = meta_fill
    cn_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Row 3: Period — label | start date (only) | blank | end date (only) ──
    ws.cell(row=3, column=1, value="Period")
    ws.cell(row=3, column=1).font      = meta_label_font
    ws.cell(row=3, column=1).fill      = meta_fill
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=3, column=2, value=s_date)
    ws.cell(row=3, column=2).number_format = DATE_FMT
    ws.cell(row=3, column=2).font      = Font(bold=True, color="1F3050", size=10)
    ws.cell(row=3, column=2).fill      = meta_fill
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=3).fill      = meta_fill   # blank separator
    ws.cell(row=3, column=4, value=e_date)
    ws.cell(row=3, column=4).number_format = DATE_FMT
    ws.cell(row=3, column=4).font      = Font(bold=True, color="1F3050", size=10)
    ws.cell(row=3, column=4).fill      = meta_fill
    ws.cell(row=3, column=4).alignment = Alignment(horizontal="center", vertical="center")
    if n_cols > 4:
        ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=n_cols)
        ws.cell(row=3, column=5).fill = meta_fill
    ws.row_dimensions[3].height = 20

    # ── Row 4: blank spacer ───────────────────────────────────────────────────
    for col in range(1, n_cols + 1):
        ws.cell(row=4, column=col).fill = meta_fill
    ws.row_dimensions[4].height = 8

    # ── Row 5: date-only reference cells — B5 = start date, D5 = end date ────
    # No labels, just the date values. Formulas reference $B$5 and $D$5.
    date_ref_fill = PatternFill("solid", start_color="D6EAF8")
    for col in range(1, n_cols + 1):
        ws.cell(row=5, column=col).fill = meta_fill
    ws.cell(row=5, column=2, value=s_date)
    ws.cell(row=5, column=2).number_format = DATE_FMT
    ws.cell(row=5, column=2).font          = Font(bold=True, color="0A2F5C", size=10)
    ws.cell(row=5, column=2).fill          = date_ref_fill
    ws.cell(row=5, column=2).alignment     = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=4, value=e_date)
    ws.cell(row=5, column=4).number_format = DATE_FMT
    ws.cell(row=5, column=4).font          = Font(bold=True, color="0A2F5C", size=10)
    ws.cell(row=5, column=4).fill          = date_ref_fill
    ws.cell(row=5, column=4).alignment     = Alignment(horizontal="center", vertical="center")
    if n_cols > 4:
        ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=n_cols)
        ws.cell(row=5, column=5).fill = meta_fill
    ws.row_dimensions[5].height = 20

    # Apply meta area outer border (rows 2-5)
    apply_outer_border(ws, 2, 5, 1, n_cols)

    # ── Row 6: blank spacer before header ────────────────────────────────────
    ws.row_dimensions[6].height = 6

    HEADER_ROW = 7
    DATA_START = HEADER_ROW + 1

    # Named cell references — $B$5 = start date, $D$5 = end date
    START_CELL = "$B$5"
    END_CELL   = "$D$5"

    FORMULA_COLS        = {"F", "H", "K", "L", "M"}
    VALUE_ONLY_COMPUTED = {"O"}

    # Build column-index map and xl() helper
    key_to_col = {k: ci for ci, k in enumerate(ordered_keys, start=1)}

    def xl(key):
        """Return the Excel column letter for an internal key."""
        return get_column_letter(key_to_col[key])

    # ── Header row (row 7) ────────────────────────────────────────────────────
    for ci, key in enumerate(ordered_keys, start=1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=f"({key}) {COLS[key]}")
        cell.font      = header_font
        cell.fill      = calc_fill if (key in FORMULA_COLS or key in VALUE_ONLY_COMPUTED) else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[HEADER_ROW].height = 40


    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), start=DATA_START):
        r = ri

        for ci, key in enumerate(ordered_keys, start=1):
            cell = ws.cell(row=r, column=ci)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")

            # ── Computed columns that get formulas ────────────────────────
            if key in FORMULA_COLS:
                cell.fill          = data_calc_fill
                cell.number_format = NUM_FMT
                cell.font          = Font(italic=True, color="1F4E79")

                if key == "F":
                    a = xl("A"); e = xl("E")
                    cell.value = (
                        f'=IF(OR(LOWER({a}{r})="software",'
                        f'LOWER({a}{r})="freehold land",'
                        f'LOWER({a}{r})="lease hold land",'
                        f'LOWER({a}{r})="leasehold land"),'
                        f'0,{e}{r}*0.05)'
                    )

                elif key == "H":
                    # Book Value as on Start Date — uses $B$4 (start date cell)
                    e = xl("E"); f_ = xl("F"); g = xl("G"); d = xl("D")
                    cell.value = (
                        f'=IF({g}{r}<=0,{e}{r},'
                        f'MAX({f_}{r},'
                        f'{e}{r}-MIN(({e}{r}-{f_}{r})/({g}{r}*12)*'
                        f'MAX(0,(YEAR({START_CELL})-YEAR({d}{r}))*12'
                        f'+(MONTH({START_CELL})-MONTH({d}{r}))'
                        f'+(DAY({START_CELL})-DAY({d}{r}))/30),'
                        f'{e}{r}-{f_}{r})))'
                    )

                elif key == "K":
                    # Period Depreciation — uses $B$4 (start) and $D$4 (end)
                    e = xl("E"); f_ = xl("F"); g = xl("G")
                    d = xl("D"); i_ = xl("I"); a = xl("A")
                    cell.value = (
                        f'=IF(OR(LOWER({a}{r})="freehold land",'
                        f'LOWER({a}{r})="lease hold land",'
                        f'LOWER({a}{r})="leasehold land"),0,'
                        f'IF({g}{r}<=0,0,'
                        f'({e}{r}-{f_}{r})/({g}{r}*12)*'
                        f'MAX(0,(YEAR(IF({i_}{r}<>"",MIN({i_}{r},{END_CELL}),{END_CELL}))'
                        f'-YEAR(MAX({d}{r},{START_CELL})))*12'
                        f'+(MONTH(IF({i_}{r}<>"",MIN({i_}{r},{END_CELL}),{END_CELL}))'
                        f'-MONTH(MAX({d}{r},{START_CELL})))'
                        f'+(DAY(IF({i_}{r}<>"",MIN({i_}{r},{END_CELL}),{END_CELL}))'
                        f'-DAY(MAX({d}{r},{START_CELL})))/30)))'
                    )

                elif key == "L":
                    e = xl("E"); o = xl("O")
                    cell.value = f'={e}{r}-{o}{r}'

                elif key == "M":
                    # Profit/Loss — uses $B$4 (start) and $D$4 (end)
                    i_ = xl("I"); n = xl("N"); l_ = xl("L")
                    cell.value = (
                        f'=IF(AND({i_}{r}<>"",{i_}{r}>={START_CELL},{i_}{r}<={END_CELL}),'
                        f'{n}{r}-{l_}{r},0)'
                    )

            # ── Accumulated Depreciation (O) – plain value, no formula ───
            elif key in VALUE_ONLY_COMPUTED:
                cell.fill          = data_calc_fill
                cell.number_format = NUM_FMT
                val = row.get(key, "")
                cell.value = "" if pd.isna(val) else val

            # ── Input columns ─────────────────────────────────────────────
            else:
                val = row.get(key, "")
                if pd.isna(val):
                    val = ""
                if isinstance(val, (pd.Timestamp, datetime)):
                    cell.value         = val.date()
                    cell.number_format = DATE_FMT
                elif key in ("D", "I") and val != "":
                    try:
                        cell.value         = pd.to_datetime(val).date()
                        cell.number_format = DATE_FMT
                    except Exception:
                        cell.value = val
                elif key in ("E", "G", "J", "N") and val != "":
                    cell.number_format = NUM_FMT
                    cell.value = val
                else:
                    cell.value = val

    # ── Total row ────────────────────────────────────────────────────────────
    total_row = DATA_START + len(df)
    tc = ws.cell(row=total_row, column=1, value="TOTAL")
    tc.font = Font(bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="1F4E79")
    tc.alignment = Alignment(horizontal="center")
    tc.border = border
    sum_cols = {"E", "F", "H", "J", "K", "L", "M", "N", "O"}
    for ci, key in enumerate(ordered_keys, start=1):
        if key in sum_cols:
            col_ltr = get_column_letter(ci)
            cell = ws.cell(
                row=total_row, column=ci,
                value=f"=SUM({col_ltr}{DATA_START}:{col_ltr}{total_row-1})"
            )
            cell.font          = Font(bold=True, color="FFFFFF")
            cell.fill          = PatternFill("solid", start_color="1F4E79")
            cell.number_format = NUM_FMT
            cell.border        = border
            cell.alignment     = Alignment(horizontal="center")

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = {"A": 22, "B": 28, "C": 12, "D": 16, "E": 16,
                  "F": 16, "G": 14, "H": 22, "I": 16, "J": 20,
                  "K": 18, "L": 18, "M": 16, "N": 16, "O": 22}
    for ci, key in enumerate(ordered_keys, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 18)

    ws.freeze_panes = f"A{DATA_START}"

    # ════════════════════════════════════════════════════════════════════════
    #  SHEET 2 – Asset Class Summary
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Asset Class Summary")
    detail_sheet = "'SLM Depreciation'"
    n_sum_cols = 7

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_sum_cols)
    t2 = ws2.cell(row=1, column=1, value="ASSET CLASS SUMMARY — SLM DEPRECIATION")
    t2.font      = title_font
    t2.fill      = title_fill
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # ── Row 2: Client Name ────────────────────────────────────────────────────
    ws2.cell(row=2, column=1, value="Client Name").font      = meta_label_font
    ws2.cell(row=2, column=1).fill                           = meta_fill
    ws2.cell(row=2, column=1).alignment                      = Alignment(horizontal="right", vertical="center")
    ws2.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_sum_cols)
    cn2 = ws2.cell(row=2, column=2,
                   value=f"='{ws.title}'!B2")   # live reference to detail sheet client name cell
    cn2.font      = Font(bold=True, color="1F3050", size=11)
    cn2.fill      = meta_fill
    cn2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[2].height = 20

    # ── Row 3: Period — label | start date (only) | blank | end date (only) ──
    ws2.cell(row=3, column=1, value="Period")
    ws2.cell(row=3, column=1).font      = meta_label_font
    ws2.cell(row=3, column=1).fill      = meta_fill
    ws2.cell(row=3, column=1).alignment = Alignment(horizontal="right", vertical="center")
    # Pull dates via formula reference from detail sheet row 5 (date-only cells)
    sd2 = ws2.cell(row=3, column=2, value=f"='{ws.title}'!$B$5")
    sd2.number_format = DATE_FMT
    sd2.font          = Font(bold=True, color="1F3050", size=10)
    sd2.fill          = meta_fill
    sd2.alignment     = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=3, column=3).fill      = meta_fill   # blank separator — no "to"
    ed2 = ws2.cell(row=3, column=4, value=f"='{ws.title}'!$D$5")
    ed2.number_format = DATE_FMT
    ed2.font          = Font(bold=True, color="1F3050", size=10)
    ed2.fill          = meta_fill
    ed2.alignment     = Alignment(horizontal="center", vertical="center")
    ws2.merge_cells(start_row=3, start_column=5, end_row=3, end_column=n_sum_cols)
    ws2.cell(row=3, column=5).fill = meta_fill
    ws2.row_dimensions[3].height = 20

    # Apply outer border on meta rows 2-3
    apply_outer_border(ws2, 2, 3, 1, n_sum_cols)

    # ── Row 4: blank spacer ───────────────────────────────────────────────────
    ws2.row_dimensions[4].height = 6

    SUM_HEADER_ROW = 5
    SUM_DATA_START = SUM_HEADER_ROW + 1

    sum_headers = [
        "Asset Class",
        "No. of Assets",
        "Cost",
        "Depreciation (Period)",
        "Closing Value",
        "Accumulated Depreciation",
        "No. of Assets Sold",
    ]
    for ci2, hdr in enumerate(sum_headers, start=1):
        is_formula_col = ci2 > 1
        cell = ws2.cell(row=SUM_HEADER_ROW, column=ci2, value=hdr)
        cell.font      = header_font
        cell.fill      = calc_fill if is_formula_col else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws2.column_dimensions[get_column_letter(ci2)].width = 26
    ws2.row_dimensions[SUM_HEADER_ROW].height = 40

    # Derive unique asset classes
    asset_classes = df["A"].dropna().astype(str).str.strip()
    asset_classes = sorted(asset_classes[asset_classes != ""].unique().tolist())

    def dcol(key):
        return get_column_letter(key_to_col[key])

    detail_last_row = DATA_START + len(df) - 1

    for row_offset, ac in enumerate(asset_classes):
        r2 = SUM_DATA_START + row_offset
        c_ac = ws2.cell(row=r2, column=1, value=ac)
        c_ac.border    = border
        c_ac.alignment = Alignment(horizontal="left", vertical="center")
        c_ac.font      = Font(bold=True, color="1F3050")

        a_range = f"{detail_sheet}!{dcol('A')}{DATA_START}:{dcol('A')}{detail_last_row}"
        e_range = f"{detail_sheet}!{dcol('E')}{DATA_START}:{dcol('E')}{detail_last_row}"
        k_range = f"{detail_sheet}!{dcol('K')}{DATA_START}:{dcol('K')}{detail_last_row}"
        l_range = f"{detail_sheet}!{dcol('L')}{DATA_START}:{dcol('L')}{detail_last_row}"
        o_range = f"{detail_sheet}!{dcol('O')}{DATA_START}:{dcol('O')}{detail_last_row}"
        i_range = f"{detail_sheet}!{dcol('I')}{DATA_START}:{dcol('I')}{detail_last_row}"
        ac_ref  = f"$A{r2}"

        formulas = [
            f"=COUNTIF({a_range},{ac_ref})",
            f"=SUMIF({a_range},{ac_ref},{e_range})",
            f"=SUMIF({a_range},{ac_ref},{k_range})",
            f"=SUMIF({a_range},{ac_ref},{l_range})",
            f"=SUMIF({a_range},{ac_ref},{o_range})",
            f'=COUNTIFS({a_range},{ac_ref},{i_range},"<>")',
        ]
        num_fmts = ["", NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, ""]

        for ci_off, (formula, nfmt) in enumerate(zip(formulas, num_fmts), start=2):
            cell = ws2.cell(row=r2, column=ci_off, value=formula)
            cell.border         = border
            cell.alignment      = Alignment(horizontal="center", vertical="center")
            cell.fill           = data_calc_fill
            cell.font           = Font(italic=True, color="1F4E79")
            if nfmt:
                cell.number_format = nfmt

    # ── Grand Total row ───────────────────────────────────────────────────────
    sum_last      = SUM_DATA_START + len(asset_classes) - 1
    sum_total_row = sum_last + 1
    gt = ws2.cell(row=sum_total_row, column=1, value="GRAND TOTAL")
    gt.font      = Font(bold=True, color="FFFFFF")
    gt.fill      = PatternFill("solid", start_color="1F4E79")
    gt.alignment = Alignment(horizontal="center")
    gt.border    = border
    for ci2 in range(2, len(sum_headers) + 1):
        col_ltr2 = get_column_letter(ci2)
        cell = ws2.cell(
            row=sum_total_row, column=ci2,
            value=f"=SUM({col_ltr2}{SUM_DATA_START}:{col_ltr2}{sum_last})"
        )
        cell.font          = Font(bold=True, color="FFFFFF")
        cell.fill          = PatternFill("solid", start_color="1F4E79")
        cell.number_format = NUM_FMT
        cell.border        = border
        cell.alignment     = Alignment(horizontal="center")



    ws2.freeze_panes = f"A{SUM_DATA_START}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 – CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
with tab_calc:
    with st.sidebar:
        st.header("⚙️ Configuration")
        client_name = st.text_input("Client Name", placeholder="Enter client name")
        start_date  = st.date_input("Start Date",  value=date(2024, 4, 1))
        end_date    = st.date_input("End Date",    value=date(2025, 3, 31))

        if start_date >= end_date:
            st.error("End Date must be after Start Date.")

        st.divider()
        uploaded_file = st.file_uploader("📂 Import Excel File", type=["xlsx","xls"])

    if uploaded_file:
        try:
            raw_df = pd.read_excel(uploaded_file, dtype=str)
            raw_df = normalise_columns(raw_df)

            for col in ["E","G","H","N"]:
                if col in raw_df.columns:
                    raw_df[col] = pd.to_numeric(raw_df[col], errors="coerce").fillna(0)
            # J (Depre as per client) is OPTIONAL – keep NaN if blank
            if "J" in raw_df.columns:
                raw_df["J"] = pd.to_numeric(raw_df["J"], errors="coerce")

            st.success(f"✅ Loaded **{len(raw_df)}** rows from the uploaded file.")

            with st.expander("📄 Preview Uploaded Data", expanded=False):
                st.dataframe(display_df(raw_df), use_container_width=True)

            if st.button("🔢 Calculate Depreciation", type="primary"):
                if start_date >= end_date:
                    st.error("Fix the date range before calculating.")
                else:
                    result = calculate_slm(raw_df, start_date, end_date)

                    st.subheader("📋 Calculation Results")
                    disp = display_df(result)

                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Total Cost",          f"₹{result['E'].sum():,.2f}")
                    col2.metric("Total Depreciation",  f"₹{result['K'].sum():,.2f}")
                    col3.metric("Total Closing Value", f"₹{result['L'].sum():,.2f}")
                    col4.metric("Total Acc. Depre.",   f"₹{result['O'].sum():,.2f}")

                    # ── Land / zero-depr info ──────────────────────────────
                    land_mask = result["A"].astype(str).str.strip().str.lower().isin(ZERO_DEPR_ASSET_CLASSES)
                    if land_mask.any():
                        land_classes = result.loc[land_mask, "A"].unique().tolist()
                        st.info(
                            f"ℹ️ **Zero Depreciation** applied to "
                            f"{land_mask.sum()} asset(s) in class(es): "
                            + ", ".join(f"*{c}*" for c in land_classes)
                        )

                    st.dataframe(disp, use_container_width=True, height=400)

                    # ── Save to history ────────────────────────────────────
                    add_history_entry(
                        client    = client_name,
                        start_dt  = start_date,
                        end_dt    = end_date,
                        num_rows  = len(result),
                        totals    = {
                            "cost":    round(float(result["E"].sum()), 2),
                            "depr":    round(float(result["K"].sum()), 2),
                            "closing": round(float(result["L"].sum()), 2),
                            "accum":   round(float(result["O"].sum()), 2),
                        },
                    )
                    st.success("📝 Result saved to history.")

                    excel_buf = to_excel(result, client_name, start_date, end_date)
                    fname = f"SLM_Depreciation_{client_name}_{end_date.strftime('%d%m%Y')}.xlsx"
                    if not client_name.strip():
                        st.warning("⚠️ Please enter a **Client Name** in the sidebar before exporting to Excel.")
                    else:
                        st.download_button(
                            label     = "📥 Export to Excel",
                            data      = excel_buf,
                            file_name = fname,
                            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

        except Exception as e:
            st.error(f"Error processing file: {e}")
            st.exception(e)

    else:
        st.info("👈 Please upload an Excel file from the sidebar to get started.")

        st.subheader("📌 Expected Column Layout")
        col_df = pd.DataFrame([
            {"Column": k, "Name": v,
             "Computed": "✅ Auto-calculated" if k in COMPUTED_COLS else "📥 Input"}
            for k, v in COLS.items()
        ])
        st.dataframe(col_df, use_container_width=True, hide_index=True)

        st.subheader("📐 Calculation Rules (VBA-aligned)")
        st.markdown("""
| Rule | Logic |
|------|-------|
| **Salvage Value (F)** | 0 if Asset Class = *Software*; else Cost × 5% |
| **Zero Depreciation** | Asset class **Freehold Land** or **Lease Hold Land** → K = 0, O = 0 |
| **Book Value (H)** | `Cost − Accumulated Depreciation up to Start Date` |
| **calcStart** | `max(Capitalized Date, Period Start)` |
| **calcEnd** | `min(Sale Date, Period End)` — or Period End if no sale |
| **Depreciation (K)** | `Monthly Depr × DateDiff(months, calcStart, calcEnd)` with fractional days: `(Day(end)−Day(start))/30` |
| **Accumulated Depr (O)** | Same fractional-month formula from Capitalization Date → calcEnd |
| **Salvage floor** | AccumDepr capped at Cost−Salvage |
| **Closing Value (L)** | `Cost − Accumulated Depreciation` |
| **Profit / Loss (M)** | `Sale Value − Closing Value` (only if sold within period) |
""")

    st.markdown("---")
    st.caption(
        f"SLM Depreciation Calculator | "
        f"User: **{st.session_state.get('username','—')}** | "
        f"Client: **{client_name if uploaded_file else '—'}** | "
        f"Period: {start_date.strftime('%d/%m/%Y')} → {end_date.strftime('%d/%m/%Y')}"
    )

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 – HISTORY
# ══════════════════════════════════════════════════════════════════════════════
with tab_hist:
    st.subheader("🕓 Calculation History")
    st.caption("Last 50 calculations across all sessions are retained.")

    history = load_history()

    hcol_a, hcol_b = st.columns([8, 2])
    with hcol_b:
        if st.button("🗑️ Clear History", type="secondary"):
            save_history([])
            st.success("History cleared.")
            st.rerun()

    if not history:
        st.info("No history yet. Run a calculation to see it here.")
    else:
        hist_df = pd.DataFrame(history)
        hist_df.rename(columns={
            "timestamp":     "Timestamp",
            "client":        "Client",
            "period_start":  "Period Start",
            "period_end":    "Period End",
            "num_assets":    "Assets",
            "total_cost":    "Total Cost (₹)",
            "total_depr":    "Total Depr (₹)",
            "total_closing": "Closing Value (₹)",
            "total_accum":   "Accum. Depr (₹)",
        }, inplace=True)

        for money_col in ["Total Cost (₹)", "Total Depr (₹)", "Closing Value (₹)", "Accum. Depr (₹)"]:
            if money_col in hist_df.columns:
                hist_df[money_col] = hist_df[money_col].apply(lambda x: f"₹{x:,.2f}")

        st.dataframe(hist_df, use_container_width=True, hide_index=True)

        # Export history to Excel
        def history_to_excel(h_df):
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb  = Workbook()
            ws  = wb.active
            ws.title = "SLM History"
            thin   = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hfill  = PatternFill("solid", start_color="1F4E79")
            hfont  = Font(bold=True, color="FFFFFF")
            cols   = list(h_df.columns)
            for ci, col in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = hfont; cell.fill = hfill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = border
            for ri, row in enumerate(h_df.itertuples(index=False), start=2):
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
            for ci in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 22
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        hist_excel = history_to_excel(hist_df)
        st.download_button(
            label     = "📥 Export History to Excel",
            data      = hist_excel,
            file_name = f"SLM_History_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
